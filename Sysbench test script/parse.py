from openpyxl.workbook import Workbook
import sys
file_name=sys.argv[1][:-4]
bp_size=file_name.split("_")[0]+"%"
num_of_threads=file_name.split("_")[1]
cpu_list=list()
device_list=list()
tps=list()
lines=list()
def write_sheet(write_wb):
	if file_name in write_wb.sheetnames:
		print("sheet update")
		write_wb.remove(write_wb[file_name])

	print("create new sheet")
	write_ws = write_wb.create_sheet(file_name)

	options=file_name.split("_")
	BP=options[0]+"%"
	num_of_thread=options[1]
	
	write_ws['B1'] ='BP size'
	write_ws['C1'] =BP
	write_ws['D1'] ='thread'
	write_ws['E1'] = int(num_of_thread)
	write_ws['F1'] = "TPS"
	write_ws['G1'] = float(tps[0])

	write_ws['B2'] = 'cpu'
	write_ws['C2'] = 'r/s'
	write_ws['D2'] = 'w/s'
	write_ws['E2'] = 'rkB/s'
	write_ws['F2'] = 'wkB/s'
	write_ws['G2'] = '%util'
	for i in range(0,len(lines)):
		row=str(3+i)
		temp=lines[i].split()
		write_ws['B'+row] = float(temp[0])
		write_ws['C'+row] = float(temp[1])
		write_ws['D'+row] = float(temp[2])
		write_ws['E'+row] = float(temp[3])
		write_ws['F'+row] = float(temp[4])
		write_ws['G'+row] = float(temp[5])
	write_ws['i1']='AVG'

	write_ws['i2']='CPU'
	write_ws['j2']='rKB/s'
	write_ws['k2']='wKB/s'
	write_ws['l2']='%util'

	write_ws['i3']='=AVERAGE(B3:B122)'
	write_ws['j3']='=AVERAGE(E3:E122)'
	write_ws['k3']='=AVERAGE(F3:F122)'
	write_ws['l3']='=AVERAGE(G3:G122)'

	write_wb.save('/home/osj/zssd_test/185G/185G.xlsx')
def line_check(line):
	if "avg-cpu" in line:
		return 1
	elif "nvme0n1" in line:
		return 2
	else:
		return 0

def abstract(file_pointer):
	file_pointer=open("/home/osj/zssd_test/185G/abs/"+file_name+"abs.txt","w+t")
	for i in range(0,len(cpu_list)):
		#if float(cpu_list[i])>=90:
		#	continue
		cpu_line=cpu_list[i]+"\t"
		try:
			device_line="\t".join(device_list[i])+"\n"
			lines.append(cpu_line+device_line)
			file_pointer.write(cpu_line+device_line)
		except(IndexError):
			continue
	file_pointer.close()
def open_file(file_pointer):

	#sysbench를 통해서 얻어지는 성능관련 txt file
	tps_pointer=open("/home/osj/zssd_test/185G/tps/tps"+file_name+".txt","r")
	while True:
		tps_line=tps_pointer.readline()
		if "transactions" in tps_line or 'per sec.' in tps_line:
			tps.append(tps_line.split()[2][1:])
			break
	
	file_pointer=open("/home/osj/zssd_test/185G/iostat/"+file_name+".txt","r")
	while True :
		line=file_pointer.readline()
		if not line:
			break
		flag=line_check(line)
		if flag==0:
			continue
		if flag==1:
			line=file_pointer.readline()
			cpu_list.append(line.split()[5])
		if flag==2:
			temp=line.split()
			device_list.append([temp[3],temp[4],temp[5],temp[6],temp[-1]])

 



if __name__ == "__main__":
	if bp_size=="%" or num_of_threads=="":
		exit()
	
	#정리되는 excel file 이름
	wb = load_workbook(filename = '/home/osj/zssd_test/185G/185G.xlsx', read_only=False, data_only=False)
	fp=None
	ret=None
	open_file(fp)
	abstract(ret)
	write_sheet(wb)
	print("BufferPool size :"+bp_size,"\nnumber of threads :"+num_of_threads,"\nwrite on file end\n")
